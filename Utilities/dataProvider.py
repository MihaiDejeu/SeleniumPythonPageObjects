import openpyxl


def get_data(sheetName):

    workbook = openpyxl.load_workbook("..//Excel//testdata.xlsx")
    sheet = workbook[sheetName]
    total_rows = sheet.max_row
    total_columns = sheet.max_column
    mainList = []

    for i in range(2, total_rows + 1):
        dataList = []
        for j in range(1, total_columns + 1):
            data = sheet.cell(row=i, column=j).value
            dataList.insert(j, data)
        mainList.insert(i, dataList)
    print(mainList)
    return mainList