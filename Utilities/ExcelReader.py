import openpyxl


def getRowCount(path, sheetName):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.max_row


def getColumnsCount(path, sheetName):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.max_column


def getCellData(path, sheetName, rowNumber, columnNumber):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.cell(row=rowNumber, column=columnNumber).value


def setCellData(path, sheetName, rowNumber, columnNumber, data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    sheet.cell(row=rowNumber, column=columnNumber).value=data
    workbook.save(path)


path = "..//excel//testdata.xlsx"
sheetName = "LoginTest"

rows = getRowCount(path, sheetName)
columns = getColumnsCount(path, sheetName)
print(rows, " and ", columns)

print(getCellData(path, sheetName, 2, 1))
setCellData(path, sheetName, 3, 3, "EDIT")
